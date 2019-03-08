import datetime
import pickle
import sys

from copy import copy
from decimal import Decimal

import openpyxl
from openpyxl.formula import Tokenizer

from fields import fields

report_date = (datetime.datetime.now() - datetime.timedelta(days=7)).strftime("%m-%d-%Y")
locale = sys.argv[1]

def write_report(pkl, sh, next_col):
    for q, entry in fields.items():
        try:
            if type(entry) == int:
                #then one-row query
                shloc = next_col + str(entry)
                write_val = pkl[q][0][0] #eg {'q1':[(100,)]}
                if write_val == None:
                    sh[shloc] = 0
                else:
                    sh[shloc] = float(write_val)
            elif type(entry) == dict: #then field is a dict of name, nrow pairs
                results = pkl[q] #eg [('APP', 273764),('微信卡包', 367695),('微信小程序', 3420), etc
                for name, datum in results: #name, value tuple
                    nrow = str(entry[name]) #name in fields dict must match query output
                    shloc = next_col + nrow
                    if datum == None:
                        sh[shloc] = 0
                    else:
                        sh[shloc] = float(datum)
        except KeyError:
            print(f'{q} is not in the results dictionary for {locale}')

def init_new_col(sh, cur_col):
    next_col = openpyxl.utils.get_column_letter(openpyxl.utils.column_index_from_string(cur_col) + 1)
    prev_col =  openpyxl.utils.get_column_letter(openpyxl.utils.column_index_from_string(cur_col) - 1)
    for i, row in enumerate(openpyxl.utils.rows_from_range('{0}2:{0}238'.format(cur_col))):
        cell = sh[row[0]]
        new_cell = sh[next_col+str(i+2)]
        new_cell._style = copy(cell._style)
        if type(cell.value) == str and '=' in cell.value:#cell contains a formula
            formula = '='+ ''.join([item.value for item in Tokenizer(cell.value).items])
            if prev_col in formula and 'SUM' not in formula:
                formula = formula.replace(cur_col, next_col)
                formula = formula.replace(prev_col, cur_col)
            elif 'SUM' in formula:
                sumpos = formula.find('SUM')
                formula = formula[:sumpos].replace(cur_col, next_col) + "SUM" + formula[sumpos+3:].replace(cur_col, next_col)
            else:
                formula = formula.replace(cur_col, next_col)
            new_cell.value = formula
        elif cell.value == 0:
            new_cell.value = 0
    #and finally, set column width
    sh.column_dimensions[next_col].width = 14.5

if __name__ == "__main__":
    with open("../temp/" + locale + ".pkl", 'rb') as pkl:
        pkl = pickle.load(pkl)
        reportpath = "../temp/" + locale + ".xlsx"
        wb = openpyxl.load_workbook(reportpath)
        sh = wb['报告－每周']
        cur_col = list(sh.columns)[-1][0].column
        next_col = openpyxl.utils.get_column_letter(openpyxl.utils.column_index_from_string(cur_col) + 1)
        init_new_col(sh, cur_col) #writes new column for next week, adding formulae and styles
        sh[next_col+str(2)] = report_date
        write_report(pkl, sh, next_col)
        wb.save(f'../temp/{locale}.xlsx')